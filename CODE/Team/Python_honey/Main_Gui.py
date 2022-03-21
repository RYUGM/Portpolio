from openpyxl import *
import requests



class Excelseat:

    def create(self,information):
        wb = Workbook()
        ws = wb.active

        ws.append(["제품 명","가격","평점","평점 등록수","제품 링크","이미지"])

        for idx,temp in enumerate(information,start=2):
            print(idx,' ',temp)
            ws.append(temp[0:])
            ws["E"+str(idx)].hyperlink = ws["E"+str(idx)].value
            ws["E"+str(idx)].value = '바로가기'
            ws["E"+str(idx)].style = "Hyperlink"

        wb.save("Coupang.xlsx")
        wb.close()

    def loadrow(self):
        rows = []
        try:
            wb = load_workbook("Coupang.xlsx")

            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                rows.append([row[0].value, row[1].value,
                        row[2].value, row[3].value,
                        row[4].value, row[5].value])
                wb.close()
            return rows
        except Exception as e:
            print(e)