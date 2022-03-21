import datetime
import sys
import os.path
import glob
import threading
from PyQt5 import uic
import requests
import re
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from bs4 import BeautifulSoup
import sys
import os, copy
import time
import pandas as pd
import smtplib
from openpyxl import *
from PyQt5.QtWidgets import *

from email.message import EmailMessage
from imap_tools import MailBox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
form_class = uic.loadUiType("assignment.ui")[0]
class MainWindow(QMainWindow,form_class):
    def __init__(self):
        super().__init__()


        self.setupUi(self)
        self.create_btn.clicked.connect(self.CreateEmail_xl)#만들기
        self.search_btn.clicked.connect(self.web_scraping)#찾기


        self.del_btn.clicked.connect(self.del_link)#삭제
        self.exit_btn.clicked.connect(QCoreApplication.instance().quit)#종료
        self.add_btn.clicked.connect(self.input_link_price)#추가


    def CreateEmail_xl(self):
        wb=Workbook()
        ws=wb.active
        ws.append(['Email','Subscribe',	'Item','CPrice','SPrice','RESULT','Name','HTML','Alter'])
        wb.save("Phase1.xlsx")
        wb.close()

    def del_link(self):
        now = datetime.datetime.now()
        print(now)
        wb = load_workbook("Phase1.xlsx")
        ws=wb.active
        link = self.lineEdit.text()
        name=self.lineEdit_3.text()
        ws.append([name,0,0,0])
        wb.save("Phase1.xlsx")
        wb.close()
        self.del_duplicate_addr1()




    def input_link_price(self):
        wb=load_workbook("Phase1.xlsx")
        ws=wb.active
        print("??")
        now = datetime.datetime.now()
        print(now)

        link=self.lineEdit.text()
        price=self.lineEdit_2.text()
        name=self.lineEdit_3.text()

        print(str(link),int(price))

        ws.append([str(name),1,str(link),int(price),'x','x'])
        print(link,price)

        wb.save("Phase1.xlsx")
        wb.close()
        self.del_duplicate_addr1()


    def del_duplicate_addr1(self):

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")#중복 날짜
        df = del_same.drop_duplicates(["Email"], keep="last")
        df.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        print("중복 고객 관리 완료")


        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 해지

        df1 = del_same[del_same['Subscribe'] == 1]  # B열에 숫자 1이 없는 행은 전부 삭제
        df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # 해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨
        print("해지 고객 관리 완료")

        # del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        # df = del_same
        # df1 = df[df['Result'] == 0]
        #
        # df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # print("최저가 알림후 삭제 완료")
        self.threading_time()
    def del_duplicate_addr2(self):

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")#중복 날짜
        df = del_same.drop_duplicates(["Email"], keep="last")
        df.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        print("중복 고객 관리 완료")


        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 해지

        df1 = del_same[del_same['Subscribe'] == 1]  # B열에 숫자 1이 없는 행은 전부 삭제
        df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # 해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨
        print("해지 고객 관리 완료")

        # del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        # df = del_same
        # df1 = df[df['Result'] == 0]
        #
        # df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # print("최저가 알림후 삭제 완료")



    def threading_time(self):
        self.tableWidget.clearContents()

        rows = self.loadrow()
        self.rowIndex = 0
        try:
            for row in rows:
                print(row)
                self.tableWidget.setItem(self.rowIndex, 0, QTableWidgetItem(row[0]))
                self.tableWidget.setItem(self.rowIndex, 1, QTableWidgetItem(row[1]))
                self.tableWidget.setItem(self.rowIndex, 2, QTableWidgetItem(row[2]))
                self.tableWidget.setItem(self.rowIndex, 3, QTableWidgetItem(row[3]))
                self.tableWidget.setItem(self.rowIndex, 4, QTableWidgetItem(row[4]))

                self.rowIndex += 1
                if (self.rowIndex > 10):
                    self.tableWidget.setRowCount(self.rowIndex + 1)
        finally:
            print("??")
            self.del_duplicate_addr2()



        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)






    def loadrow(self):
        rows = []
        try:
            wb = load_workbook("Phase1.xlsx")

            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                rows.append([row[0].value,
                        str(row[3].value), str(row[4].value),
                        str(row[5].value), row[2].value])
                wb.close()
            return rows
        except Exception as e:
            print(e)





    def web_scraping(self):
        wb = load_workbook("Phase1.xlsx")
        ws = wb.active

        try:
            for temp in ws.iter_rows(min_row=2):

                url = temp[2].value
                print(url)

                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36"}
                res = requests.get(url, headers=headers)
                res.raise_for_status()

                soup = BeautifulSoup(res.text, "lxml")

                name = soup.select("h2.prod-buy-header__title")

                price = soup.find("span", attrs={"class": "total-price"}).get_text()
                # sold_out = soup.find("span", attrs={"class": "prod-not-find-known__buy__info__txt"}).get_text()

                Myname = name[0].text
                price = (price.replace(',', '').replace('원', ''))

                result = (int(price) - int(temp[3].value))
                print("result=?", result)
                ws.append([temp[0].value, temp[1].value, temp[2].value, temp[3].value, price, result, Myname])
                # if sold_out is not None:
                #     print(sold_out)
                #     ws.append([temp[0].value, temp[1].value, temp[2].value, temp[3].value, '품절', '품절', Myname])
                #




                if (result < 0):
                    ws.append([temp[0].value, temp[1].value, temp[2].value, temp[3].value, price, result, Myname])



        finally:
            print("최저가 검색결과")
            wb.save("Phase1.xlsx")
            wb.close()
            self.del_duplicate_addr1()






if __name__ == "__main__":
    app=QApplication(sys.argv)

    myWindow = MainWindow()

    myWindow.show()
    app.exec_()