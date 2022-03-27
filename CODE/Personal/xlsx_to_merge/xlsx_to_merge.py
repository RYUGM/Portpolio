import sys
import os.path
import glob
import easygui
import openpyxl.chart
from PyQt5.QtWidgets import *
from PyQt5 import uic, QtGui, QtWidgets, QtCore
from PyQt5.QtWidgets import *
from openpyxl import *
from PyQt5.QtWidgets import QTableWidgetItem
from openpyxl import *
from openpyxl.chart import LineChart, Reference
from openpyxl.chart import BarChart, Reference
import addexcel

from easygui import *
from PyQt5.QtGui import *
from PyQt5.QtCore import pyqtSlot
from datetime import datetime

form_class = uic.loadUiType("assignment.ui")[0]


class MyMainApp(QMainWindow, form_class):
    def __init__(self):
        super().__init__()

        self.row = 0
        self.col = 0
        self.value = 1
        self.rowIndex = 0
        self.ax = addexcel
        self.save_Folder = ""
        self.save_File = []
        self.setupUi(self)

        # self.btn1.clicked.connect(self.savefn)
        # self.btn2.clicked.connect(self.loadfn)
        # self.btn3.clicked.connect(self.createfn)
        # self.btn4.clicked.connect(self.barMakefile)
        self.addbtn.clicked.connect(self.tool_btn_slot)
        self.tbtn.clicked.connect(self.tool2_btn_slot)
        self.selbtn.clicked.connect(self.Addforfiles)

        # self.stox.clicked.connect(self.s_to_x_btn)
        # self.stox2.clicked.connect(self.x_to_s_btn)

        # self.grbox = QtWidgets.QGroupBox(self)
        # self.listView = QtWidgets.QListWidget(self)
        # self.grbox.raise_()

    @pyqtSlot()
    def tool_btn_slot(self): #폴더에서 선택한 엑셀 파일 표기
        path = easygui.fileopenbox(multiple=True)
        self.save_File = path
        model = QStandardItemModel()

        for p in path:
            model.appendRow(QStandardItem(p))
        self.listView.setModel(model)

    @pyqtSlot()
    def tool2_btn_slot(self): #병합 결과물 저장 경로 표기
        self.save_Folder = easygui.diropenbox() + "\\result.xlsx"
        self.lineEdit.setText((self.save_Folder))

    @pyqtSlot()
    def s_to_x_btn(self): #gsheet 확장자를 xlsx로 변환하면 그냥 바뀌는줄...이건 실패 
        path = self.dirline.text()

        # apath = glob.glob(path+"/*.xlsx")
        # bpath = glob.glob(path+"/*.gsheet")


        files = glob.glob(path+"/*.gsheet")
        # if files == apath:
        for x in files:
                if not os.path.isdir(x):
                    files = os.path.splitext(x)
                # try:
                    os.rename(x, files[0] + ".xlsx")
                # except:
                #     pass
    # def S_to_X(self,path):
    @pyqtSlot()
    def x_to_s_btn(self): #이것도 마찬가지로 실패...
        path = self.dirline.text()
        # path = self.save_Folder
        # apath = glob.glob(path+"/*.xlsx")
        # bpath = glob.glob(path+"/*.gsheet")

        files = glob.glob(path+"/*.xlsx")
        # if files == apath:
        for x in files:
            if not os.path.isdir(x):
                files = os.path.splitext(x)
                # try:
                os.rename(x, files[0] + ".gsheet")


    def Addforfiles(self): #선택한 엑셀파일 병합하는 함수

        input_wb = Workbook()
        input_ws = input_wb.active
        for file in self.save_File:
            wb = load_workbook(filename=file, data_only=True)
            print(file)
            ws = wb.active
            for row in ws.iter_rows():
                # print("data",row)
                data = []
                for cell in row:
                    data.append(cell.value)
                    print("cell", cell)
                input_ws.append(data)
            input_ws.append(["---------",datetime.today(),file,"---------"]) #병합했을때 엑셀 파일간 구분을 짓기 위한 표시
        input_wb.save(self.save_Folder)

    # def data_input(self, ws):
    #     input_wb = Workbook()
    #     input_ws = input_wb.active
    #
    #     for row in ws.iter_rows():
    #
    #         # print("data",row)
    #         data = []
    #
    #         for cell in row:
    #             data.append(cell.value)
    #             print("cell", cell)
    #
    #         input_ws.append(data)
    #
    #     input_wb.save(self.save_Folder)

    # print(results)
    # for i in results:
    #     input_wb = Workbook()
    #     input_ws = input_wb.active
    #     input_ws.append(i.value)
    #     print("i",i)
    # input_wb.save(self.save_Folder)

    # wb.save(self.save_Folder)
    # input_wb = Workbook()
    # input_ws = input_wb.active
    # for i in results:
    #     input_ws.append(i)
    #     print("i=",i)
    # input_wb.save(self.save_Folder)

    # for row in ws.iter_rows():
    #     data = []
    #     print(data)
    #     for cell in row:
    #         data.append(cell.value)
    #         input_ws.append(data)
    # input_wb.save(self.save_Folder)
    # self.data_input(ws, input_ws)

    # def data_input(self,ws,input_ws):
    #     for row in ws.iter_rows():
    #         data = []
    #         for cell in row:
    #             data.append(cell.value)
    #         input_ws.append(data)

    # Addforfiles(self.save_Folder)
    def savefn(self):   #파기
        wb = load_workbook("a.xlsx")
        ws = wb.active

        kor = self.textbox1.text()
        eng = self.textbox2.text()
        math = self.textbox3.text()
        tot = int(kor) + int(eng) + int(math)
        avg = int(tot / 3 * 100) / 100
        print('tot = ', tot)
        print('avg = ', avg)
        ws.append([int(kor), int(eng), int(math), tot, avg])
        wb.save('a.xlsx')
        wb.close()

    def loadfn(self): #파기

        kor = self.textbox1.text()
        eng = self.textbox2.text()
        math = self.textbox3.text()
        tot = int(kor) + int(eng) + int(math)
        avg = int(tot / 3 * 100) / 100
        lst = [str(kor), str(eng), str(math), str(tot), str(avg)]
        for i in lst:

            item = QTableWidgetItem()

            item.setText(i)
            self.table1.setItem(self.row, self.col, item)

            self.value += 1
            self.col += 1
            if (self.col == 5):
                self.row += 1
                self.col = 0
            if (self.row > 9):
                self.table1.setRowCount(self.row + 1)

    # def loadfile1(self):
    #     rows = []
    #
    #
    #     wb = load_workbook('sample_a.xlsx')
    #
    #     ws = wb.active
    #
    #     for row in ws.iter_rows(min_row=2):
    #         rows.append([row[0].value, row[1].value,
    #                      row[2].value, row[3].value,
    #                      row[4].value])
    #         self.tabl1.setItem(self.rowIndex, 0, QTableWidgetItem(row[0]))
    #         self.tabl1.setItem(self.rowIndex, 1, QTableWidgetItem(row[1]))
    #         self.tabl1.setItem(self.rowIndex, 2, QTableWidgetItem(row[2]))
    #         self.tabl1.setItem(self.rowIndex, 3, QTableWidgetItem(row[3]))
    #         self.tabl1.setItem(self.rowIndex, 4, QTableWidgetItem(row[4]))
    #         self.rowIndex += 1
    #         wb.close()
    #
    #     print(rows)
    #

    #
    #

    #     print("Add")
    #     # files = os.scandir(path)
    #
    #     files = os.listdir(path)
    #     input_wb = Workbook()
    #     input_ws = input_wb.active
    #     for file in files:
    #         wb = load_workbook(path + "\\" + file)
    #         ws = wb.active
    #
    #         data_input(ws, input_ws)
    #     input_wb.save(path + "\\" + "Add_file.xlsx")
    #
    # def data_input(ws, input_ws):
    #         for row in ws.iter_rows():
    #             data = []
    #             for cell in row:
    #                 data.append(cell.value)
    #             input_ws.append(data)
    #             # return input_ws
    #
    def Add_excel(self): #파기
        pass

    def createfn(self): #파기
        wb = Workbook()
        ws = wb.active
        ws.append(['국어', '영어', '수학', '총점', '평균'])
        wb.save('a.xlsx')
        wb.close()

    def barMakefile(self): #파기
        wb = load_workbook("a.xlsx")
        ws = wb['Sheet']

        # bar_value = Reference(ws,min_col=4,min_row=1, max_col=5,max_row=10)
        # ws, min_col=1, min_row=2, max_col=3, max_row=10
        bar_value = Reference(ws, min_col=1, min_row=1, max_col=5, max_row=20)
        bar_chart = BarChart()

        bar_chart.add_data(bar_value, from_rows=False, titles_from_data=True)
        # cate = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_col=5, max_row=20)

        # 왜 안돼
        # bar_chart.set_categories(cate)
        ws.add_chart(bar_chart, "G10")

        wb.save("sample_a.xlsx")
        wb.close()

    #     wb = load_workbook("a.xlsx")
    #     ws = wb.active
    #
    #     bar_value=openpyxl.chart.reference(ws,1,1,10)
    #     bar_chart=openpyxl.chart.BarChart()
    #     bar_chart.add_data(bar_value)
    #     ws.add_chart(bar_chart,"G10")
    #     wb.save("a_bar.xlsx")
    #     wb.close()
    # barMakefile()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyMainApp()
    myWindow.show()

    app.exec_()
