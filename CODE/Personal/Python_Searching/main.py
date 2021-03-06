import datetime
import sys
import os.path
import glob
import threading
from PyQt5 import uic
import requests
import re
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from bs4 import BeautifulSoup
import sys
import os, copy

import pandas as pd
import smtplib
from openpyxl import *
from PyQt5.QtWidgets import *
from account_PC import *
from email.message import EmailMessage
from imap_tools import MailBox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import schedule
form_class = uic.loadUiType("assignment.ui")[0]
class MainWindow(QMainWindow,form_class):
    def __init__(self):
        super().__init__()

        self.setupUi(self)
        self.create_btn.clicked.connect(self.CreateEmail_xl)
        self.recive_btn.clicked.connect(self.recive_mail)
        self.exit_btn.clicked.connect(QCoreApplication.instance().quit)


    def CreateEmail_xl(self): #데이터 저장에 필요한 엑셀 파일 생성 함수
        wb=Workbook()
        ws=wb.active
        ws.append(["Email","Subscribe","Item","CPrice","SPrice","RESULT","Name","HTML","Alter"])
        wb.save("Phase1.xlsx")
        wb.close()

        wb1=Workbook()
        ws1=wb1.active
        ws1.append(["Email","Subscribe","Item","CPrice","SPrice","RESULT","Name","HTML","Alter"])
        wb1.save("Phase2.xlsx") #Phase2는 이력 남기기용 (언제 어떤상품을 등록했고 얼마일때 구입했다 등의...)
        wb1.close()
     
    def del_duplicate_addr1(self): #중복 고객 삭제 함수 

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")#중복 이메일
        df = del_same.drop_duplicates(["Email"], keep="last")
        df.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        print("중복 고객 관리 완료")


        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 해지

        df1 = del_same[del_same['Subscribe'] == 1]  # B열에 숫자 1이 없는 행은 전부 삭제
        df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # 해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨
        print("해지 고객 관리 완료")

        # del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        # df = del_same
        # df1 = df[df['Result'] == 0]
        #
        # df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # print("최저가 알림후 삭제 완료")
        self.web_scraping()

    def del_duplicate_addr2(self): 

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        df = del_same.drop_duplicates(["Email"], keep="last")
        df.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        print("중복 고객 관리 완료")

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 해지

        df1 = del_same[del_same['Subscribe'] == 1]  # B열에 숫자 1이 없는 행은 전부 삭제
        df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # 해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨
        print("해지 고객 관리 완료")

        # del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        # df = del_same
        # df1 = df[df['Result'] == 0]
        #
        # df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # print("최저가 알림후 삭제 완료")
        self.send_mail()


    def del_duplicate_addr3(self): #똑같은 중복 삭제 함수가 세개나 있는 이유는 
        #이메일 체크 함수 -> 중복/해지 함수1 -> 웹스크래핑 -> 중복/해지 함수2 -> 
        # 이메일 발송 함수(최저가 유무에따라) -> 중복/해지 함수3 -> 일정시간 대기하는 함수 -> 이메일 체크 
        #순서로 반복하기때문에 같은 기능을 하더라고 하나의 중복/해지 함수로는 각 단계에 따른 함수를 호출하기 힘들었기 때문

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        df = del_same.drop_duplicates(["Email"], keep="last")
        df.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        print("중복 고객 관리 완료")

        del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 해지

        df1 = del_same[del_same['Subscribe'] == 1]  # B열에 숫자 1이 없는 행은 전부 삭제
        df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # 해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨
        print("해지 고객 관리 완료")

        # del_same = pd.read_excel("Phase1.xlsx", sheet_name="Sheet")  # 중복 이메일
        # df = del_same
        # df1 = df[df['Result'] == 0]
        #
        # df1.to_excel("Phase1.xlsx", sheet_name="Sheet", index=False)
        # print("최저가 알림후 삭제 완료")
        self.threading_time()

    def threading_time(self): #웹 스크래핑 결과를 UI에 출력하고 일정시간 대기하는 함수
        wb=load_workbook("Phase1.xlsx")
        ws=wb.active
        list_view=[]
        now = datetime.datetime.now()

        for listv in ws.values:
           list_view.append([listv[0],listv[3],listv[4],listv[5],now])
        model = QStandardItemModel()
        for text in list_view:
            print(text)
            model.appendRow(QStandardItem(str(text).replace('[','').replace(']','').replace('.datetime','')))
        self.listView.setModel(model)

        wb.close()

        now = datetime.datetime.now()
        print("현재시간 ",now)
        threading.Timer(120, self.recive_mail).start()





    def recive_mail(self):#이메일 받기

        wb=load_workbook("Phase1.xlsx")
        ws=wb.active

        mailbox = MailBox("imap.gmail.com", 993)
        mailbox.login(EMAIL_ADDRESS, EMAIL_PASSWORD, initial_folder="INBOX")
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

        try:
            for msg in mailbox.fetch(('UNSEEN TEXT "AI04-RYU"'), limit=5, reverse=False):

                if msg.from_ is not None:
                    if "해지" in msg.text:  # 내용에 해지란 단어가 있으면 0을 추가
                        print("해지")

                        ws.append([msg.from_, 0])

                        print("해지 고객", msg.from_)

                        cu_address = msg.from_  # 해지 안내 메일 발송
                        msg = EmailMessage()
                        msg["Subject"] = cu_address + "님 해지완료. "
                        msg["From"] = EMAIL_ADDRESS
                        msg["TO"] = cu_address
                        msg.set_content(cu_address + "고객님 정상적으로 해지 완료되었습니다. \n이용해 주셔서 감사합니다.")

                        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
                            smtp.ehlo()
                            smtp.starttls()
                            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

                            smtp.send_message(msg)
                        wb.save("Phase1.xlsx")
                        wb.close()


                    else:

                        user_price = msg.text

                        user_price1 = user_price.replace("AI04-RYU", "")
                        # user_price1 = re.findall(r'\d+', user_price)
                        print("msgtext",msg.text)
                        print("0번째", user_price1)

                        user_item = msg.subject

                        ws.append([msg.from_, 1, user_item, user_price1,1,1,1,1,1])

                        print("고객 메일주소={0},희망 아이템 ={1}, 희망 가격{2}(이하)".format(msg.from_, user_item, user_price1))


        finally:
             time.sleep(5)


        wb.save("Phase1.xlsx")
        wb.close()
        self.del_duplicate_addr1()









    def send_mail(self): #류경문 - 이메일 보내기


        wb = load_workbook("Phase1.xlsx")#현재 저장된 고객 명단
        ws = wb.active
        wb1=load_workbook("Phase2.xlsx")
        ws1=wb1.active
        try:
            for temp in ws.iter_rows(min_row=2):
                mail_check = temp[5].value #temp[5] 에는 (현재 가격 - 설정가격) 의 결과가 저장되어있음
                print(temp[0].value, temp[1].value, temp[2].value, temp[3].value, temp[4].value, temp[5].value,
                      temp[6].value, temp[7].value, temp[8].value)
                print("최저가 메일 보내기")
                if (mail_check < 0): #(현재 가격 - 설정가격) 이 0보다 작다는건 실제 판매가격이 고객이 설정한 가격보다 낮다는 뜻
                    ws.append([temp[0].value, 0, temp[2].value, temp[3].value, temp[4].value, temp[5].value,
                      temp[6].value, temp[7].value, temp[8].value])
                    ws1.append([temp[0].value,temp[1].value , temp[2].value, temp[3].value, temp[4].value, temp[5].value,
                      temp[6].value, temp[7].value, temp[8].value])

                    s = smtplib.SMTP('smtp.gmail.com', 587)
                    s.starttls()
                    s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
                    msg = MIMEMultipart('alternative')
                    msg['Subject'] = "최저가 알리미에서" + temp[0].value + "고객님에게 전달해 드립니다!"
                    msg['From'] = EMAIL_ADDRESS
                    msg['To'] = temp[0].value

                    resultNum = int(temp[5].value) * -1
                    print("result", resultNum)

                    body = '''\n고객님께서 관심가지신 상품이 희망하신 가격 보다 <h2>''' + str(resultNum) + '''원</h2> 내려갔습니다.'지금  바로 확인해보세요!\n\n
                                ''' + temp[7].value + '''
                                \n\n\n<a href="''' + temp[2].value + '''">\n\n구입을 원하시면 여기를 클릭~!</a>\n\n\n

                        \n\n\n더이상 최저가 알림을 받기 싫으신 고객님 께서는 \n
                        답장 버튼을 누르셔서 바로 보내기를 눌러주시면 구독 해지 처리 해 드리겠습니다.\n
                          AI04-RYU 프로그램을 사용해 주셔서 감사합니다\n'''

                    body = (body.replace('[', '').replace(',', '').replace(']', ''))
                    text = MIMEText(body, 'html', 'utf-8')
                    msg.attach(text)

                    wb.save("Phase1.xlsx")
                    wb.close()
                    s.sendmail(EMAIL_ADDRESS, temp[0].value, msg.as_string())
                    s.quit()
                    print(temp[0].value + "님에게 발송 완료")

                else:
                    print("최저가 없음")
        finally:
            print("진짜 최저가 없음")


        wb1.save("Phase2.xlsx")
        wb1.close()
        wb.save("Phase1.xlsx")
        wb.close()
        self.del_duplicate_addr3()

    def web_scraping(self): #웹 스크래핑 함수
        wb = load_workbook("Phase1.xlsx")
        ws = wb.active


        for temp in ws.iter_rows(min_row=2):


            url = temp[2].value
            print(url)

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36"}
            res = requests.get(url, headers=headers)
            res.raise_for_status()

            soup = BeautifulSoup(res.text, "lxml")

            name = soup.select("h2.prod-buy-header__title")

            price = soup.find("span", attrs={"class": "total-price"}).get_text()

            price = (price.replace(',', '').replace('원', ''))
            Myname = name[0].text
            html_img_tag = soup.select("img.prod-image__detail")
            html1_tag = soup.select("div.prod-price-onetime")


            result = (int(price) - int(temp[3].value))
            print("result=?", result)

            if (result < 0):

                ws.append([temp[0].value,temp[1].value ,temp[2].value, temp[3].value,price,result,Myname,str(html_img_tag+html1_tag)])
        print("최저가 검색결과")
        wb.save("Phase1.xlsx")
        wb.close()
        self.del_duplicate_addr2()





if __name__ == "__main__":
    app=QApplication(sys.argv)


    myWindow = MainWindow()
    myWindow.show()
    app.exec_()

